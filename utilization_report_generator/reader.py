"""
Core logic for reading source data and processing utilization records
"""

import os
from collections import defaultdict
import openpyxl

from .config import (
    WORKING_DAYS, SHEET_CONFIG, DATA_START_ROW,
    COL_ASSOC_ID, COL_ASSOC_NAME, COL_GRADE, COL_PROJECT,
    COL_ACCOUNT, COL_BILLABILITY, COL_COUNTRY, COL_OO, COL_CITY,
    UTIL_HIGH, UTIL_MEDIUM
)


def safe_float(v, default=0.0):
    """Safely convert value to float"""
    try:
        return float(v) if v is not None else default
    except (ValueError, TypeError):
        return default


def avg_util(record, months_order):
    """Calculate average utilization across months"""
    vals = [record['months'][m]['util'] for m in months_order if m in record['months']]
    return round(sum(vals) / len(vals), 1) if vals else 0.0


def read_source_data(src_path):
    """
    Read and process employee data from source Excel file.
    
    Returns:
        tuple: (records list, months_order list, statistics dict)
    """
    if not os.path.exists(src_path):
        raise FileNotFoundError(f"Source file not found: {src_path}")
    
    print(f"[1/4] Reading source file: {src_path}")
    
    wb = openpyxl.load_workbook(src_path, data_only=True)
    
    # Detect which months are actually present
    months_order = []
    for _, month_configs in SHEET_CONFIG:
        for month_name, _, _ in month_configs:
            if month_name not in months_order and month_name in WORKING_DAYS:
                months_order.append(month_name)
    
    all_records = {}
    
    # Read data from each sheet
    for sheet_name, month_configs in SHEET_CONFIG:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found — skipping.")
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        for row in rows[DATA_START_ROW:]:
            if not any(row):
                continue
            
            assoc_id = row[COL_ASSOC_ID] if COL_ASSOC_ID < len(row) else None
            if not assoc_id:
                continue
            
            # Create or update record
            if assoc_id not in all_records:
                all_records[assoc_id] = {
                    'Associate ID': assoc_id,
                    'Associate Name': row[COL_ASSOC_NAME] if COL_ASSOC_NAME < len(row) else '',
                    'Grade': row[COL_GRADE] if COL_GRADE < len(row) else '',
                    'Account': row[COL_ACCOUNT] if COL_ACCOUNT < len(row) else '',
                    'Project': row[COL_PROJECT] if COL_PROJECT < len(row) else '',
                    'Billability': row[COL_BILLABILITY] if COL_BILLABILITY < len(row) else '',
                    'Country': row[COL_COUNTRY] if COL_COUNTRY < len(row) else '',
                    'Onsite/Offshore': row[COL_OO] if COL_OO < len(row) else '',
                    'City': row[COL_CITY] if COL_CITY < len(row) else '',
                    'months': {}
                }
            
            # Process each month's data
            for month_name, fc_idx, act_idx in month_configs:
                wd = WORKING_DAYS.get(month_name, 22)
                forecast = row[fc_idx] if fc_idx < len(row) else None
                actual = row[act_idx] if act_idx < len(row) else None
                
                leave = safe_float(actual) if actual is not None else safe_float(forecast)
                available = wd - leave
                util_pct = round((available / wd) * 100, 1) if wd > 0 else 0.0
                
                all_records[assoc_id]['months'][month_name] = {
                    'wd': wd,
                    'forecast': safe_float(forecast),
                    'actual': safe_float(actual) if actual is not None else None,
                    'leave': leave,
                    'available': available,
                    'util': util_pct,
                }
    
    records = list(all_records.values())
    print(f"  -> {len(records)} associates loaded across {len(months_order)} months.")
    
    # Calculate statistics
    avg_utils = [avg_util(r, months_order) for r in records]
    
    stats = {
        'total_associates': len(records),
        'total_forecast': sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) 
                              for r in records for m in months_order),
        'total_actual': sum(safe_float(r['months'].get(m, {}).get('actual', 0)) 
                           for r in records for m in months_order),
        'overall_avg': round(sum(avg_utils) / len(avg_utils), 1) if avg_utils else 0,
        'low_count': sum(1 for u in avg_utils if u < UTIL_MEDIUM),
        'med_count': sum(1 for u in avg_utils if UTIL_MEDIUM <= u < UTIL_HIGH),
        'high_count': sum(1 for u in avg_utils if u >= UTIL_HIGH),
    }
    
    # Add monthly statistics
    monthly_stats = []
    for m in months_order:
        wd = WORKING_DAYS.get(m, 22)
        mf = round(sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) 
                       for r in records), 1)
        ma = round(sum(safe_float(r['months'].get(m, {}).get('actual', 0)) 
                       for r in records), 1)
        ml = round(sum(r['months'].get(m, {}).get('leave', 0) for r in records), 1)
        mu = round(sum(r['months'].get(m, {}).get('util', 0) for r in records) / len(records), 1) \
            if records else 0
        
        monthly_stats.append({
            'month': m,
            'wd': wd,
            'forecast': mf,
            'actual': ma,
            'leave': ml,
            'util': mu
        })
    
    stats['monthly_stats'] = monthly_stats
    
    return records, months_order, stats
