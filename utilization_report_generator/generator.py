"""
Main Excel and HTML report generators
"""

import os
from datetime import date
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import WORKING_DAYS, UTIL_HIGH, UTIL_MEDIUM
from .reader import safe_float, avg_util


def fill(hex_color):
    """Create a fill with hex color"""
    return PatternFill("solid", fgColor=hex_color)


def xfont(bold=False, color="595959", size=10, italic=False):
    """Create a font with parameters"""
    return Font(bold=bold, color=color, size=size, italic=italic)


def border_thin():
    """Create thin borders"""
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def border_medium():
    """Create medium borders"""
    s = Side(style="medium", color="595959")
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    """Center alignment"""
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left_align():
    """Left alignment"""
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def util_color(u, light=False):
    """Get color based on utilization percentage"""
    if u >= UTIL_HIGH:
        return "C6EFCE" if light else "22c55e"
    elif u >= UTIL_MEDIUM:
        return "FFF0B3" if light else "f59e0b"
    else:
        return "FFB3B3" if light else "ef4444"


def util_bg(u):
    """Get background color based on utilization"""
    if u >= UTIL_HIGH:
        return "E2EFDA"
    elif u >= UTIL_MEDIUM:
        return "FFF2CC"
    else:
        return "FFE0E0"


def util_label(u):
    """Get label for utilization level"""
    if u >= UTIL_HIGH:
        return "High (>=90%)"
    elif u >= UTIL_MEDIUM:
        return "Medium (80-90%)"
    else:
        return "Low (<80%)"


def generate_excel_report(records, months_order, out_path):
    """Generate comprehensive Excel report with multiple sheets"""
    print(f"\n[2/4] Generating Excel report...")
    
    C_DARK_BLUE = "1F3864"
    C_MID_BLUE = "2E75B6"
    C_VERY_LIGHT = "DEEAF1"
    C_WHITE = "FFFFFF"
    C_GREY = "F2F2F2"
    MONTH_FILLS = ["2E75B6", "1B6CA8", "155A8A", "0F4A72"]
    
    wb = openpyxl.Workbook()
    
    def write_title(ws, text, cols, row=1, bg=C_DARK_BLUE, size=13, height=30):
        ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
        ws[f"A{row}"] = text
        ws[f"A{row}"].font = Font(bold=True, size=size, color="FFFFFF")
        ws[f"A{row}"].fill = fill(bg)
        ws[f"A{row}"].alignment = center()
        ws.row_dimensions[row].height = height
    
    def write_header_cell(ws, row, col, val, bg=C_DARK_BLUE):
        c = ws.cell(row, col, val)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border_thin()
        return c
    
    def write_data_cell(ws, row, col, val, bg="FFFFFF", bold=False, align="center"):
        c = ws.cell(row, col, val)
        c.fill = fill(bg)
        c.font = xfont(size=9, bold=bold)
        c.alignment = center() if align == "center" else left_align()
        c.border = border_thin()
        return c
    
    # ── Sheet 1: Dashboard ──────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_view.showGridLines = False
    
    write_title(ws_dash, "QEA - UHG Leave & Utilization Report 2026 (H1)", 14, row=1, size=16, height=36)
    
    ws_dash.merge_cells("A2:N2")
    ws_dash["A2"] = "Utilization % = (Working Days - Leave Days) / Working Days x 100   |   Leave Used = Actuals if available, else Forecast"
    ws_dash["A2"].font = Font(italic=True, size=9, color="FFFFFF")
    ws_dash["A2"].fill = fill(C_MID_BLUE)
    ws_dash["A2"].alignment = center()
    ws_dash.row_dimensions[2].height = 18
    
    avg_utils = [avg_util(r, months_order) for r in records]
    total_assoc = len(records)
    tot_forecast = sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) for r in records for m in months_order)
    tot_actual = sum(safe_float(r['months'].get(m, {}).get('actual', 0)) for r in records for m in months_order)
    overall_avg = round(sum(avg_utils) / len(avg_utils), 1) if avg_utils else 0
    low_count = sum(1 for u in avg_utils if u < UTIL_MEDIUM)
    med_count = sum(1 for u in avg_utils if UTIL_MEDIUM <= u < UTIL_HIGH)
    high_count = sum(1 for u in avg_utils if u >= UTIL_HIGH)
    
    kpis = [
        ("Total Associates", total_assoc, C_MID_BLUE),
        ("H1 Forecast (Days)", int(tot_forecast), C_MID_BLUE),
        ("H1 Actuals (Days)", int(tot_actual), C_MID_BLUE),
        ("Avg H1 Util %", f"{overall_avg}%", C_MID_BLUE),
        ("High Util (>=90%)", high_count, "C6EFCE"),
        ("Medium Util (80-90%)", med_count, "FFF0B3"),
        ("Low Util (<80%)", low_count, "FFB3B3"),
    ]
    
    col_pairs = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12), (13, 14)]
    for (lbl, val, kf), (c1, c2) in zip(kpis, col_pairs):
        lc1, lc2 = get_column_letter(c1), get_column_letter(c2)
        ws_dash.merge_cells(f"{lc1}4:{lc2}4")
        ws_dash.merge_cells(f"{lc1}5:{lc2}5")
        ws_dash[f"{lc1}4"] = lbl
        ws_dash[f"{lc1}4"].font = Font(bold=True, size=9, color="595959")
        ws_dash[f"{lc1}4"].fill = fill(C_GREY)
        ws_dash[f"{lc1}4"].alignment = center()
        ws_dash[f"{lc1}5"] = val
        ws_dash[f"{lc1}5"].font = Font(bold=True, size=14, color=C_DARK_BLUE)
        ws_dash[f"{lc1}5"].fill = fill(kf)
        ws_dash[f"{lc1}5"].alignment = center()
    
    ws_dash.row_dimensions[4].height = 20
    ws_dash.row_dimensions[5].height = 32
    
    ws_dash.merge_cells("A7:H7")
    ws_dash["A7"] = "Monthly Leave & Utilization Summary"
    ws_dash["A7"].font = Font(bold=True, size=11, color="FFFFFF")
    ws_dash["A7"].fill = fill(C_DARK_BLUE)
    ws_dash["A7"].alignment = center()
    ws_dash.row_dimensions[7].height = 22
    
    for ci, h in enumerate(["Month", "Working Days", "Forecast (Days)", "Actuals (Days)", "Leave Used", "Available Days", "Avg Util %", "Associates"], 1):
        write_header_cell(ws_dash, 8, ci, h, C_MID_BLUE)
    ws_dash.row_dimensions[8].height = 22
    
    for ri, m in enumerate(months_order, 9):
        n = len([r for r in records if m in r['months']])
        wd = WORKING_DAYS.get(m, 22)
        mf = round(sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) for r in records), 1)
        ma = round(sum(safe_float(r['months'].get(m, {}).get('actual', 0)) for r in records), 1)
        ml = round(sum(r['months'].get(m, {}).get('leave', 0) for r in records), 1)
        mav = round(sum(r['months'].get(m, {}).get('available', 0) for r in records), 1)
        mu = round(sum(r['months'].get(m, {}).get('util', 0) for r in records) / len(records), 1) if records else 0
        rf = C_VERY_LIGHT if ri % 2 == 0 else C_WHITE
        for ci, val in enumerate([m, wd, mf, ma, ml, mav, f"{mu}%", n], 1):
            write_data_cell(ws_dash, ri, ci, val, bg=rf)
        ws_dash.row_dimensions[ri].height = 18
    
    for col, w in zip("ABCDEFGHIJKLMN", [14, 14, 16, 16, 14, 14, 12, 12, 10, 10, 10, 10, 10, 10]):
        ws_dash.column_dimensions[col].width = w
    
    # ── Sheet 2: Employee Utilization ───────────────────────
    ws_emp = wb.create_sheet("Employee Utilization")
    ws_emp.sheet_view.showGridLines = False
    ws_emp.freeze_panes = "J3"
    
    static_hdrs = ["S.No", "Associate ID", "Associate Name", "Grade", "Account", "Project",
                   "Billability", "Country", "Onsite/Offshore", "City", "Avg H1 Util %", "Alert"]
    n_static = len(static_hdrs)
    month_subs = ["Forecast\n(Days)", "Actuals\n(Days)", "Leave\nUsed", "Available\nDays", "Util %"]
    
    total_cols = n_static + len(months_order) * len(month_subs)
    write_title(ws_emp, "Employee-Level Leave & Utilization Detail - H1 2026", total_cols, row=1, height=30)
    
    for ci, h in enumerate(static_hdrs, 1):
        write_header_cell(ws_emp, 2, ci, h)
    
    for mi, m in enumerate(months_order):
        start_ci = n_static + 1 + mi * len(month_subs)
        end_ci = start_ci + len(month_subs) - 1
        mfill = MONTH_FILLS[mi % len(MONTH_FILLS)]
        ws_emp.merge_cells(f"{get_column_letter(start_ci)}2:{get_column_letter(end_ci)}2")
        c = ws_emp.cell(2, start_ci, m)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = fill(mfill)
        c.alignment = center()
        c.border = border_medium()
        for si, sh in enumerate(month_subs):
            write_header_cell(ws_emp, 3, start_ci + si, sh, mfill)
    
    ws_emp.row_dimensions[2].height = 22
    ws_emp.row_dimensions[3].height = 30
    
    for ri, rec in enumerate(sorted(records, key=lambda r: r['Associate Name']), 4):
        avg_u = avg_util(rec, months_order)
        rf = util_bg(avg_u)
        vals = [ri - 3, rec['Associate ID'], rec['Associate Name'], rec['Grade'],
                rec['Account'], rec['Project'], rec['Billability'],
                rec['Country'], rec['Onsite/Offshore'], rec['City'],
                f"{avg_u}%", util_label(avg_u)]
        for ci, v in enumerate(vals, 1):
            write_data_cell(ws_emp, ri, ci, v, bg=rf, bold=(ci == 3), align="left" if ci == 3 else "center")
        
        for mi, m in enumerate(months_order):
            md = rec['months'].get(m, {})
            u = md.get('util', '')
            mvals = [md.get('forecast', ''), md.get('actual', ''),
                    md.get('leave', ''), md.get('available', ''),
                    f"{u}%" if u != '' else '']
            for si, v in enumerate(mvals):
                ci = n_static + 1 + mi * len(month_subs) + si
                cf = util_color(u, light=True) if si == 4 and u != '' else rf
                write_data_cell(ws_emp, ri, ci, v, bg=cf, bold=(si == 4))
        ws_emp.row_dimensions[ri].height = 16
    
    for ci, w in enumerate([5, 12, 24, 7, 22, 28, 10, 14, 12, 12, 10, 16], 1):
        ws_emp.column_dimensions[get_column_letter(ci)].width = w
    
    wb.save(out_path)
    print(f"  -> Excel saved: {out_path}")
    
    return avg_utils, overall_avg, low_count, med_count, high_count, tot_forecast, tot_actual


def generate_html_dashboard(records, months_order, out_path, overall_avg, low_count, med_count, high_count, tot_forecast, tot_actual):
    """Generate HTML dashboard with charts and tables"""
    print(f"\n[3/4] Generating HTML dashboard...")
    
    total_assoc = len(records)
    
    # Generate monthly statistics
    monthly_stats = []
    for m in months_order:
        n = len([r for r in records if m in r['months']])
        mf = round(sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) for r in records), 1)
        ma = round(sum(safe_float(r['months'].get(m, {}).get('actual', 0)) for r in records), 1)
        ml = round(sum(r['months'].get(m, {}).get('leave', 0) for r in records), 1)
        mu = round(sum(r['months'].get(m, {}).get('util', 0) for r in records) / n, 1) if n else 0
        monthly_stats.append({'month': m, 'forecast': mf, 'actual': ma, 'leave': ml, 'util': mu})
    
    # Account rollup
    acct_data = defaultdict(list)
    for r in records:
        acct_data[r['Account']].append(r)
    
    acct_rollup = []
    for acct, recs in sorted(acct_data.items(), key=lambda x: -len(x[1])):
        n = len(recs)
        avg_u = round(sum(avg_util(r, months_order) for r in recs) / n, 1)
        month_utils = {m: round(sum(r['months'].get(m, {}).get('util', 0) for r in recs) / n, 1) for m in months_order}
        acct_rollup.append({'account': acct, 'count': n, 'avg_util': avg_u, 'months': month_utils})
    
    # Grade rollup
    grade_data = defaultdict(list)
    for r in records:
        grade_data[r['Grade']].append(r)
    grade_order = ['A', 'SA', 'M', 'SM', 'AD', 'PA', 'PAT', 'Cont']
    all_grades = grade_order + [g for g in grade_data if g not in grade_order]
    grade_rollup = [{'grade': g, 'count': len(grade_data[g]),
                    'avg_util': round(sum(avg_util(r, months_order) for r in grade_data[g]) / len(grade_data[g]), 1)}
                   for g in all_grades if g in grade_data]
    
    # Country rollup
    country_data = defaultdict(list)
    for r in records:
        country_data[r['Country']].append(r)
    country_rollup = [{'country': c, 'count': len(recs),
                      'avg_util': round(sum(avg_util(r, months_order) for r in recs) / len(recs), 1)}
                     for c, recs in sorted(country_data.items(), key=lambda x: -len(x[1]))]
    
    # Low util employees
    low_emp = sorted([r for r in records if avg_util(r, months_order) < UTIL_HIGH],
                    key=lambda r: avg_util(r, months_order))[:20]
    
    def badge(u):
        cls = "badge-green" if u >= UTIL_HIGH else "badge-yellow" if u >= UTIL_MEDIUM else "badge-red"
        return f'<span class="badge {cls}">{u}%</span>'
    
    def td_util(u):
        c = "#dcfce7" if u >= UTIL_HIGH else "#fef9c3" if u >= UTIL_MEDIUM else "#fee2e2"
        return f'<td style="background:{c};text-align:center;font-weight:700;">{u}%</td>'
    
    acct_rows = ""
    for i, a in enumerate(acct_rollup):
        mcells = "".join(td_util(a['months'][m]) for m in months_order)
        acct_rows += f"<tr><td>{i+1}</td><td style='font-weight:600'>{a['account']}</td><td style='text-align:center'>{a['count']}</td><td style='text-align:center'>{badge(a['avg_util'])}</td>{mcells}</tr>"
    
    low_rows = ""
    for i, r in enumerate(low_emp):
        avg_u = avg_util(r, months_order)
        mcells = "".join(td_util(r['months'].get(m, {}).get('util', 0)) for m in months_order)
        low_rows += f"<tr><td>{i+1}</td><td style='font-weight:600'>{r['Associate Name']}</td><td style='text-align:center'>{r['Grade']}</td><td>{r['Account'][:30]}</td><td style='text-align:center'>{r['Country']}</td><td style='text-align:center'>{badge(avg_u)}</td>{mcells}</tr>"
    
    monthly_rows = ""
    for ms in monthly_stats:
        c = "#dcfce7" if ms['util'] >= UTIL_HIGH else "#fef9c3" if ms['util'] >= UTIL_MEDIUM else "#fee2e2"
        monthly_rows += f"<tr><td style='font-weight:700'>{ms['month']}</td><td style='text-align:center'>{WORKING_DAYS.get(ms['month'], '-')}</td><td style='text-align:center'>{ms['forecast']}</td><td style='text-align:center'>{ms['actual']}</td><td style='text-align:center'>{ms['leave']}</td><td style='background:{c};text-align:center;font-weight:700'>{ms['util']}%</td></tr>"
    
    month_headers = "".join(f"<th>{m}</th>" for m in months_order)
    today = date.today().strftime("%d %B %Y")
    
    # Basic HTML template
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>QEA UHG Utilization Dashboard 2026</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',sans-serif;background:#f0f4f8;color:#1e293b}}
.header{{background:linear-gradient(135deg,#1e3a5f,#2e75b6);color:#fff;padding:28px 36px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 4px 12px rgba(0,0,0,.2)}}
.header h1{{font-size:1.55rem;font-weight:700}}
.container{{max-width:1400px;margin:0 auto;padding:24px}}
.section-title{{font-size:1rem;font-weight:700;color:#1e3a5f;margin:28px 0 14px;padding-left:12px;border-left:4px solid #2e75b6}}
.kpi-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:14px;margin-bottom:8px}}
.kpi-card{{background:#fff;border-radius:12px;padding:18px 14px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.07);border-top:4px solid #2e75b6}}
.kpi-card .val{{font-size:1.8rem;font-weight:800;color:#1e3a5f}}
.kpi-card .lbl{{font-size:.7rem;color:#64748b;margin-top:4px}}
.table-card{{background:#fff;border-radius:12px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.07);margin-bottom:20px;overflow-x:auto}}
.table-card h3{{font-size:.92rem;font-weight:700;color:#1e3a5f;margin-bottom:14px}}
table{{width:100%;border-collapse:collapse;font-size:.81rem}}
th{{background:#1e3a5f;color:#fff;padding:10px;text-align:left;font-size:.77rem}}
td{{padding:8px 10px;border-bottom:1px solid #e2e8f0}}
.badge{{display:inline-block;padding:3px 10px;border-radius:20px;font-size:.74rem;font-weight:700}}
.badge-green{{background:#dcfce7;color:#16a34a}}
.badge-yellow{{background:#fef9c3;color:#b45309}}
.badge-red{{background:#fee2e2;color:#dc2626}}
.footer{{text-align:center;padding:20px;color:#94a3b8;font-size:.77rem}}
</style>
</head>
<body>
<div class="header">
  <h1>QEA – UHG Leave & Utilization Dashboard 2026 (H1)</h1>
  <div style="text-align:right;font-size:.85rem;opacity:.8"><div>Generated: {today}</div></div>
</div>
<div class="container">
  <div class="section-title">Key Performance Indicators</div>
  <div class="kpi-grid">
    <div class="kpi-card"><div class="val">{total_assoc}</div><div class="lbl">Total Associates</div></div>
    <div class="kpi-card"><div class="val">{int(tot_forecast)}</div><div class="lbl">H1 Forecast Leave (Days)</div></div>
    <div class="kpi-card"><div class="val">{int(tot_actual)}</div><div class="lbl">H1 Actual Leave (Days)</div></div>
    <div class="kpi-card"><div class="val" style="color:#2e75b6">{overall_avg}%</div><div class="lbl">Avg H1 Utilization</div></div>
    <div class="kpi-card" style="border-top-color:#22c55e"><div class="val" style="color:#22c55e">{high_count}</div><div class="lbl">High Util (>={UTIL_HIGH}%)</div></div>
    <div class="kpi-card" style="border-top-color:#f59e0b"><div class="val" style="color:#f59e0b">{med_count}</div><div class="lbl">Medium Util ({UTIL_MEDIUM}-{UTIL_HIGH-1}%)</div></div>
    <div class="kpi-card" style="border-top-color:#ef4444"><div class="val" style="color:#ef4444">{low_count}</div><div class="lbl">Low Util (<{UTIL_MEDIUM}%)</div></div>
  </div>

  <div class="section-title">Monthly Summary</div>
  <div class="table-card">
    <table><thead><tr><th>Month</th><th>Working Days</th><th>Forecast (Days)</th><th>Actuals (Days)</th><th>Leave Used</th><th>Avg Util %</th></tr></thead>
    <tbody>{monthly_rows}</tbody></table>
  </div>

  <div class="section-title">Account-Level Utilization Rollup</div>
  <div class="table-card">
    <table><thead><tr><th>#</th><th>Account</th><th>Associates</th><th>Avg H1 Util%</th>{month_headers}</tr></thead>
    <tbody>{acct_rows}</tbody></table>
  </div>

  <div class="section-title">Attention List (Top 20 Lowest Utilization)</div>
  <div class="table-card">
    <table><thead><tr><th>#</th><th>Associate Name</th><th>Grade</th><th>Account</th><th>Country</th><th>Avg Util%</th>{month_headers}</tr></thead>
    <tbody>{low_rows}</tbody></table>
  </div>

  <div class="footer">QEA – UHG Leave & Utilization Report &nbsp;•&nbsp; Generated on {today}</div>
</div>
</body></html>"""
    
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  -> HTML saved: {out_path}")
