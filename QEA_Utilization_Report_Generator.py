"""
========================================================================
QEA – UHG Leave & Utilization Report Generator
========================================================================
Usage:
    python QEA_Utilization_Report_Generator.py
    -> You will be prompted to enter the path to the source Excel file.

Or pass the file path directly:
    python QEA_Utilization_Report_Generator.py "C:/path/to/file.xlsx"

Output:
    - QEA-UHG-Utilization-Report-<YYYY-MM-DD>.xlsx   (Excel report)
    - QEA-UHG-Utilization-Dashboard-<YYYY-MM-DD>.html (HTML dashboard)
    Both files are saved in the same folder as the source file.

Requirements:
    pip install openpyxl

Configuration (edit section below if months/structure changes):
    - WORKING_DAYS      : working days per month
    - SHEET_CONFIG      : sheet names and column index mapping
    - UTIL_THRESHOLDS   : green/yellow/red cutoffs
========================================================================
"""

import sys
import os
from datetime import date
from collections import defaultdict

# ── Install openpyxl if missing ──────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION — Edit here when uploading a new file
# ============================================================

# Working days per month (adjust for public holidays if needed)
WORKING_DAYS = {
    "Jan'26": 21, "Feb'26": 20, "Mar'26": 22,
    "Apr'26": 22, "May'26": 21, "Jun'26": 22,
    "Jul'26": 23, "Aug'26": 21, "Sep'26": 22,
    "Oct'26": 22, "Nov'26": 21, "Dec'26": 23,
}

# Sheet configuration:
# Each entry: (sheet_name, [(month_name, forecast_col_index, actual_col_index), ...])
# Column indices are 0-based (col A = 0)
SHEET_CONFIG = [
    ('2026-Jan-Feb-Mar', [
        ("Jan'26", 14, 15),
        ("Feb'26", 18, 19),
        ("Mar'26", 22, 23),
    ]),
    ('2026-Apr-May-Jun', [
        ("Apr'26", 14, 15),
        ("May'26", 18, 19),
        ("Jun'26", 22, 23),
    ]),
]

# Utilization thresholds
UTIL_HIGH   = 90   # >= HIGH  -> Green
UTIL_MEDIUM = 80   # >= MEDIUM -> Yellow, else Red

# Data row start index (0-based): row 0 = month header, row 1 = column header, row 2+ = data
DATA_START_ROW = 2

# Column indices for employee info (0-based)
COL_ASSOC_ID    = 1
COL_ASSOC_NAME  = 2
COL_GRADE       = 3
COL_PROJECT     = 5
COL_ACCOUNT     = 7
COL_BILLABILITY = 10
COL_COUNTRY     = 11
COL_OO          = 12   # Onsite/Offshore
COL_CITY        = 13

# ============================================================
# HELPERS
# ============================================================

def safe_float(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except:
        return default

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def xfont(bold=False, color="595959", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="595959")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def util_color(u, light=False):
    if u >= UTIL_HIGH:   return "C6EFCE" if light else "22c55e"
    elif u >= UTIL_MEDIUM: return "FFF0B3" if light else "f59e0b"
    else:                  return "FFB3B3" if light else "ef4444"

def util_bg(u):
    if u >= UTIL_HIGH:   return "E2EFDA"
    elif u >= UTIL_MEDIUM: return "FFF2CC"
    else:                  return "FFE0E0"

def util_label(u):
    if u >= UTIL_HIGH:   return "High (>=90%)"
    elif u >= UTIL_MEDIUM: return "Medium (80-90%)"
    else:                  return "Low (<80%)"

def avg_util(rec, months_order):
    vals = [rec['months'][m]['util'] for m in months_order if m in rec['months']]
    return round(sum(vals)/len(vals), 1) if vals else 0.0

# ============================================================
# STEP 1: READ SOURCE DATA
# ============================================================

def read_source(src_path):
    print(f"\n[1/4] Reading source file: {src_path}")
    wb = openpyxl.load_workbook(src_path, data_only=True)

    # Detect which months are actually present
    months_order = []
    for _, month_configs in SHEET_CONFIG:
        for month_name, _, _ in month_configs:
            if month_name not in months_order and month_name in WORKING_DAYS:
                months_order.append(month_name)

    all_records = {}

    for sheet_name, month_configs in SHEET_CONFIG:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found — skipping.")
            continue
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[DATA_START_ROW:]:
            if not any(row):
                continue
            assoc_id = row[COL_ASSOC_ID]
            if not assoc_id:
                continue

            if assoc_id not in all_records:
                all_records[assoc_id] = {
                    'Associate ID':    assoc_id,
                    'Associate Name':  row[COL_ASSOC_NAME]  or '',
                    'Grade':           row[COL_GRADE]        or '',
                    'Account':         row[COL_ACCOUNT]      or '',
                    'Project':         row[COL_PROJECT]      or '',
                    'Billability':     row[COL_BILLABILITY]  or '',
                    'Country':         row[COL_COUNTRY]      or '',
                    'Onsite/Offshore': row[COL_OO]           or '',
                    'City':            row[COL_CITY]         or '',
                    'months': {}
                }

            for month_name, fc_idx, act_idx in month_configs:
                wd       = WORKING_DAYS.get(month_name, 22)
                forecast = row[fc_idx]  if fc_idx  < len(row) else None
                actual   = row[act_idx] if act_idx < len(row) else None
                leave    = safe_float(actual) if actual is not None else safe_float(forecast)
                available = wd - leave
                util_pct  = round((available / wd) * 100, 1) if wd > 0 else 0.0
                all_records[assoc_id]['months'][month_name] = {
                    'wd':       wd,
                    'forecast': safe_float(forecast),
                    'actual':   safe_float(actual) if actual is not None else None,
                    'leave':    leave,
                    'available': available,
                    'util':     util_pct,
                }

    records = list(all_records.values())
    print(f"  -> {len(records)} associates loaded across {len(months_order)} months.")
    return records, months_order

# ============================================================
# STEP 2: GENERATE EXCEL REPORT
# ============================================================

def generate_excel(records, months_order, out_path):
    print(f"\n[2/4] Generating Excel report...")

    C_DARK_BLUE  = "1F3864"
    C_MID_BLUE   = "2E75B6"
    C_VERY_LIGHT = "DEEAF1"
    C_WHITE      = "FFFFFF"
    C_GREY       = "F2F2F2"
    MONTH_FILLS  = ["2E75B6","1B6CA8","155A8A","0F4A72","0A3A5A","052A42"]

    wb = openpyxl.Workbook()

    def write_title(ws, text, cols, row=1, bg=C_DARK_BLUE, size=13, height=30):
        ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
        ws[f"A{row}"] = text
        ws[f"A{row}"].font = Font(bold=True, size=size, color="FFFFFF")
        ws[f"A{row}"].fill = fill(bg)
        ws[f"A{row}"].alignment = center()
        ws.row_dimensions[row].height = height

    def write_header_cell(ws, row, col, val, bg=C_DARK_BLUE):
        c = ws.cell(row, col, val)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border_thin()
        return c

    def write_data_cell(ws, row, col, val, bg="FFFFFF", bold=False, align="center"):
        c = ws.cell(row, col, val)
        c.fill = fill(bg)
        c.font = xfont(size=9, bold=bold)
        c.alignment = center() if align == "center" else left_align()
        c.border = border_thin()
        return c

    # ── Sheet 1: Dashboard ──────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_view.showGridLines = False

    write_title(ws_dash, "QEA - UHG Leave & Utilization Report 2026 (H1)", 14, row=1, size=16, height=36)

    ws_dash.merge_cells("A2:N2")
    ws_dash["A2"] = "Utilization % = (Working Days - Leave Days) / Working Days x 100   |   Leave Used = Actuals if available, else Forecast"
    ws_dash["A2"].font      = Font(italic=True, size=9, color="FFFFFF")
    ws_dash["A2"].fill      = fill(C_MID_BLUE)
    ws_dash["A2"].alignment = center()
    ws_dash.row_dimensions[2].height = 18

    avg_utils   = [avg_util(r, months_order) for r in records]
    total_assoc = len(records)
    tot_forecast= sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) for r in records for m in months_order)
    tot_actual  = sum(safe_float(r['months'].get(m, {}).get('actual',   0)) for r in records for m in months_order)
    overall_avg = round(sum(avg_utils)/len(avg_utils), 1) if avg_utils else 0
    low_count   = sum(1 for u in avg_utils if u < UTIL_MEDIUM)
    med_count   = sum(1 for u in avg_utils if UTIL_MEDIUM <= u < UTIL_HIGH)
    high_count  = sum(1 for u in avg_utils if u >= UTIL_HIGH)

    kpis = [
        ("Total Associates",      total_assoc,        C_MID_BLUE),
        ("H1 Forecast (Days)",    int(tot_forecast),  C_MID_BLUE),
        ("H1 Actuals (Days)",     int(tot_actual),    C_MID_BLUE),
        ("Avg H1 Util %",         f"{overall_avg}%",  C_MID_BLUE),
        ("High Util (>=90%)",     high_count,         "C6EFCE"),
        ("Medium Util (80-90%)",  med_count,          "FFF0B3"),
        ("Low Util (<80%)",       low_count,          "FFB3B3"),
    ]
    col_pairs = [(1,2),(3,4),(5,6),(7,8),(9,10),(11,12),(13,14)]
    for (lbl, val, kf), (c1, c2) in zip(kpis, col_pairs):
        lc1, lc2 = get_column_letter(c1), get_column_letter(c2)
        ws_dash.merge_cells(f"{lc1}4:{lc2}4")
        ws_dash.merge_cells(f"{lc1}5:{lc2}5")
        ws_dash[f"{lc1}4"] = lbl
        ws_dash[f"{lc1}4"].font      = Font(bold=True, size=9, color="595959")
        ws_dash[f"{lc1}4"].fill      = fill(C_GREY)
        ws_dash[f"{lc1}4"].alignment = center()
        ws_dash[f"{lc1}5"] = val
        ws_dash[f"{lc1}5"].font      = Font(bold=True, size=14, color=C_DARK_BLUE)
        ws_dash[f"{lc1}5"].fill      = fill(kf)
        ws_dash[f"{lc1}5"].alignment = center()
    ws_dash.row_dimensions[4].height = 20
    ws_dash.row_dimensions[5].height = 32

    ws_dash.merge_cells("A7:H7")
    ws_dash["A7"] = "Monthly Leave & Utilization Summary"
    ws_dash["A7"].font      = Font(bold=True, size=11, color="FFFFFF")
    ws_dash["A7"].fill      = fill(C_DARK_BLUE)
    ws_dash["A7"].alignment = center()
    ws_dash.row_dimensions[7].height = 22

    for ci, h in enumerate(["Month","Working Days","Forecast (Days)","Actuals (Days)","Leave Used","Available Days","Avg Util %","Associates"], 1):
        write_header_cell(ws_dash, 8, ci, h, C_MID_BLUE)
    ws_dash.row_dimensions[8].height = 22

    for ri, m in enumerate(months_order, 9):
        n   = len([r for r in records if m in r['months']])
        wd  = WORKING_DAYS.get(m, 22)
        mf  = round(sum(safe_float(r['months'].get(m,{}).get('forecast', 0)) for r in records), 1)
        ma  = round(sum(safe_float(r['months'].get(m,{}).get('actual',   0)) for r in records), 1)
        ml  = round(sum(r['months'].get(m,{}).get('leave', 0) for r in records), 1)
        mav = round(sum(r['months'].get(m,{}).get('available', 0) for r in records), 1)
        mu  = round(sum(r['months'].get(m,{}).get('util', 0) for r in records) / len(records), 1) if records else 0
        rf  = C_VERY_LIGHT if ri % 2 == 0 else C_WHITE
        for ci, val in enumerate([m, wd, mf, ma, ml, mav, f"{mu}%", n], 1):
            write_data_cell(ws_dash, ri, ci, val, bg=rf)
        ws_dash.row_dimensions[ri].height = 18

    for col, w in zip("ABCDEFGHIJKLMN", [14,14,16,16,14,14,12,12,10,10,10,10,10,10]):
        ws_dash.column_dimensions[col].width = w

    # ── Sheet 2: Employee Utilization ───────────────────────
    ws_emp = wb.create_sheet("Employee Utilization")
    ws_emp.sheet_view.showGridLines = False
    ws_emp.freeze_panes = "J3"

    static_hdrs  = ["S.No","Associate ID","Associate Name","Grade","Account","Project",
                    "Billability","Country","Onsite/Offshore","City","Avg H1 Util %","Alert"]
    n_static     = len(static_hdrs)
    month_subs   = ["Forecast\n(Days)","Actuals\n(Days)","Leave\nUsed","Available\nDays","Util %"]

    total_cols = n_static + len(months_order) * len(month_subs)
    write_title(ws_emp, "Employee-Level Leave & Utilization Detail - H1 2026", total_cols, row=1, height=30)

    for ci, h in enumerate(static_hdrs, 1):
        write_header_cell(ws_emp, 2, ci, h)
    for ci in range(1, n_static+1):
        ws_emp.cell(3, ci, "").fill   = fill(C_DARK_BLUE)
        ws_emp.cell(3, ci).border = border_thin()

    for mi, m in enumerate(months_order):
        start_ci = n_static + 1 + mi * len(month_subs)
        end_ci   = start_ci + len(month_subs) - 1
        mfill    = MONTH_FILLS[mi % len(MONTH_FILLS)]
        ws_emp.merge_cells(f"{get_column_letter(start_ci)}2:{get_column_letter(end_ci)}2")
        c = ws_emp.cell(2, start_ci, m)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = fill(mfill); c.alignment = center(); c.border = border_medium()
        for si, sh in enumerate(month_subs):
            write_header_cell(ws_emp, 3, start_ci+si, sh, mfill)

    ws_emp.row_dimensions[2].height = 22
    ws_emp.row_dimensions[3].height = 30

    for ri, rec in enumerate(sorted(records, key=lambda r: r['Associate Name']), 4):
        avg_u = avg_util(rec, months_order)
        rf    = util_bg(avg_u)
        vals  = [ri-3, rec['Associate ID'], rec['Associate Name'], rec['Grade'],
                 rec['Account'], rec['Project'], rec['Billability'],
                 rec['Country'], rec['Onsite/Offshore'], rec['City'],
                 f"{avg_u}%", util_label(avg_u)]
        for ci, v in enumerate(vals, 1):
            write_data_cell(ws_emp, ri, ci, v, bg=rf, bold=(ci==3), align="left" if ci==3 else "center")

        for mi, m in enumerate(months_order):
            md   = rec['months'].get(m, {})
            u    = md.get('util', '')
            mvals= [md.get('forecast',''), md.get('actual',''),
                    md.get('leave',''), md.get('available',''),
                    f"{u}%" if u != '' else '']
            for si, v in enumerate(mvals):
                ci = n_static + 1 + mi * len(month_subs) + si
                cf = util_color(u, light=True) if si == 4 and u != '' else rf
                write_data_cell(ws_emp, ri, ci, v, bg=cf, bold=(si==4))
        ws_emp.row_dimensions[ri].height = 16

    for ci, w in enumerate([5,12,24,7,22,28,10,14,12,12,10,16], 1):
        ws_emp.column_dimensions[get_column_letter(ci)].width = w
    for mi in range(len(months_order)):
        for si in range(len(month_subs)):
            ci = n_static + 1 + mi * len(month_subs) + si
            ws_emp.column_dimensions[get_column_letter(ci)].width = 10

    # ── Sheet 3: Low Util Alerts ────────────────────────────
    ws_alert = wb.create_sheet("Low Util Alerts")
    ws_alert.sheet_view.showGridLines = False
    ws_alert.freeze_panes = "A4"

    total_alert_cols = 8 + len(months_order)
    write_title(ws_alert, f"Low Utilization Alert Report - Associates with Avg H1 Util < {UTIL_MEDIUM}%",
                total_alert_cols, row=1, bg="C00000", height=30)
    ws_alert.merge_cells(f"A2:{get_column_letter(total_alert_cols)}2")
    ws_alert["A2"] = f"Associates below {UTIL_MEDIUM}% average H1 utilization - sorted by lowest utilization first."
    ws_alert["A2"].font = Font(italic=True, size=9, color="C00000")
    ws_alert["A2"].fill = fill("FFE0E0"); ws_alert["A2"].alignment = center()
    ws_alert.row_dimensions[2].height = 18

    alert_hdrs = ["S.No","Associate ID","Associate Name","Grade","Account",
                  "Country","Onsite/Offshore","Avg Util %"] + [f"{m}\nUtil %" for m in months_order]
    for ci, h in enumerate(alert_hdrs, 1):
        write_header_cell(ws_alert, 3, ci, h, "C00000")
    ws_alert.row_dimensions[3].height = 28

    low_recs = sorted([r for r in records if avg_util(r, months_order) < UTIL_MEDIUM], key=lambda r: avg_util(r, months_order))
    for ri, rec in enumerate(low_recs, 4):
        avg_u = avg_util(rec, months_order)
        rf    = "FFD7D7" if avg_u < 70 else "FFE0E0"
        base  = [ri-3, rec['Associate ID'], rec['Associate Name'], rec['Grade'],
                 rec['Account'], rec['Country'], rec['Onsite/Offshore'], f"{avg_u}%"]
        month_cells = [f"{rec['months'].get(m,{}).get('util','')}%" if rec['months'].get(m,{}).get('util','') != '' else '' for m in months_order]
        for ci, v in enumerate(base + month_cells, 1):
            c = write_data_cell(ws_alert, ri, ci, v, bg=rf, bold=(ci==8), align="left" if ci==3 else "center")
            if ci >= 9:
                try:
                    uu = float(str(v).replace('%',''))
                    c.fill = fill(util_color(uu, light=True))
                except: pass
        ws_alert.row_dimensions[ri].height = 16

    for ci, w in enumerate([5,12,24,7,22,14,12,10]+[10]*len(months_order), 1):
        ws_alert.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 4: Account Rollup ─────────────────────────────
    ws_acct = wb.create_sheet("Account Rollup")
    ws_acct.sheet_view.showGridLines = False
    ws_acct.freeze_panes = "A3"

    acct_total_cols = 5 + len(months_order) * 3
    write_title(ws_acct, "Account-Level Leave & Utilization Rollup - H1 2026", acct_total_cols, row=1, height=30)

    acct_hdrs = ["S.No","Account Name","# Associates","Avg H1 Util %","Alert"]
    for m in months_order: acct_hdrs += [f"{m}\nForecast", f"{m}\nActuals", f"{m}\nAvg Util %"]
    for ci, h in enumerate(acct_hdrs, 1):
        write_header_cell(ws_acct, 2, ci, h)
    ws_acct.row_dimensions[2].height = 30

    acct_data = defaultdict(list)
    for r in records: acct_data[r['Account']].append(r)

    for ri, (acct, recs) in enumerate(sorted(acct_data.items(), key=lambda x: -len(x[1])), 3):
        n     = len(recs)
        avg_u = round(sum(avg_util(r, months_order) for r in recs)/n, 1)
        rf    = util_bg(avg_u)
        for ci, v in enumerate([ri-2, acct, n, f"{avg_u}%", util_label(avg_u)], 1):
            write_data_cell(ws_acct, ri, ci, v, bg=rf, bold=(ci in [2,4]), align="left" if ci==2 else "center")
        for mi, m in enumerate(months_order):
            mf = round(sum(safe_float(r['months'].get(m,{}).get('forecast',0)) for r in recs), 1)
            ma = round(sum(safe_float(r['months'].get(m,{}).get('actual',  0)) for r in recs), 1)
            mu = round(sum(r['months'].get(m,{}).get('util',0) for r in recs)/n, 1)
            ci_base = 6 + mi * 3
            for ci, v in zip([ci_base, ci_base+1, ci_base+2], [mf, ma, f"{mu}%"]):
                cf = util_color(mu, light=True) if ci == ci_base+2 else rf
                write_data_cell(ws_acct, ri, ci, v, bg=cf, bold=(ci==ci_base+2))
        ws_acct.row_dimensions[ri].height = 18

    for ci, w in enumerate([5,26,10,10,14]+[11]*len(months_order)*3, 1):
        ws_acct.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 5: Grade Rollup ───────────────────────────────
    ws_grade = wb.create_sheet("Grade Rollup")
    ws_grade.sheet_view.showGridLines = False

    grade_total_cols = 4 + len(months_order) * 3
    write_title(ws_grade, "Grade-Level Leave & Utilization Rollup - H1 2026", grade_total_cols, row=1, height=30)

    grade_hdrs = ["Grade","# Associates","Avg H1 Util %","Alert"]
    for m in months_order: grade_hdrs += [f"{m}\nForecast", f"{m}\nActuals", f"{m}\nAvg Util %"]
    for ci, h in enumerate(grade_hdrs, 1):
        write_header_cell(ws_grade, 2, ci, h)
    ws_grade.row_dimensions[2].height = 30

    grade_data  = defaultdict(list)
    for r in records: grade_data[r['Grade']].append(r)
    grade_order = ['A','SA','M','SM','AD','PA','PAT','Cont']
    all_grades  = grade_order + [g for g in grade_data if g not in grade_order]

    for ri, grade in enumerate([g for g in all_grades if g in grade_data], 3):
        recs  = grade_data[grade]
        n     = len(recs)
        avg_u = round(sum(avg_util(r, months_order) for r in recs)/n, 1)
        rf    = util_bg(avg_u)
        for ci, v in enumerate([grade, n, f"{avg_u}%", util_label(avg_u)], 1):
            write_data_cell(ws_grade, ri, ci, v, bg=rf, bold=True)
        for mi, m in enumerate(months_order):
            mf = round(sum(safe_float(r['months'].get(m,{}).get('forecast',0)) for r in recs), 1)
            ma = round(sum(safe_float(r['months'].get(m,{}).get('actual',  0)) for r in recs), 1)
            mu = round(sum(r['months'].get(m,{}).get('util',0) for r in recs)/n, 1)
            ci_base = 5 + mi * 3
            for ci, v in zip([ci_base, ci_base+1, ci_base+2], [mf, ma, f"{mu}%"]):
                cf = util_color(mu, light=True) if ci == ci_base+2 else rf
                write_data_cell(ws_grade, ri, ci, v, bg=cf, bold=(ci==ci_base+2))
        ws_grade.row_dimensions[ri].height = 18

    for ci, w in enumerate([10,12,12,14]+[11]*len(months_order)*3, 1):
        ws_grade.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 6: Country Rollup ─────────────────────────────
    ws_country = wb.create_sheet("Country Rollup")
    ws_country.sheet_view.showGridLines = False

    country_total_cols = 5 + len(months_order)
    write_title(ws_country, "Country & Onsite/Offshore - Utilization Rollup - H1 2026", country_total_cols, row=1, height=30)

    country_hdrs = ["Country","Onsite/Offshore","# Associates","Avg H1 Util %","Alert"] + [f"{m}\nAvg Util %" for m in months_order]
    for ci, h in enumerate(country_hdrs, 1):
        write_header_cell(ws_country, 2, ci, h)
    ws_country.row_dimensions[2].height = 30

    country_oo_data = defaultdict(list)
    for r in records: country_oo_data[(r['Country'], r['Onsite/Offshore'])].append(r)

    for ri, ((country, oo), recs) in enumerate(sorted(country_oo_data.items(), key=lambda x: -len(x[1])), 3):
        n     = len(recs)
        avg_u = round(sum(avg_util(r, months_order) for r in recs)/n, 1)
        rf    = util_bg(avg_u)
        for ci, v in enumerate([country, oo, n, f"{avg_u}%", util_label(avg_u)], 1):
            write_data_cell(ws_country, ri, ci, v, bg=rf, bold=True)
        for mi, m in enumerate(months_order):
            mu = round(sum(r['months'].get(m,{}).get('util',0) for r in recs)/n, 1)
            write_data_cell(ws_country, ri, 6+mi, f"{mu}%", bg=util_color(mu, light=True), bold=True)
        ws_country.row_dimensions[ri].height = 18

    for ci, w in enumerate([16,14,12,12,14]+[12]*len(months_order), 1):
        ws_country.column_dimensions[get_column_letter(ci)].width = w

    wb.save(out_path)
    print(f"  -> Excel saved: {out_path}")
    return avg_utils, overall_avg, low_count, med_count, high_count, tot_forecast, tot_actual

# ============================================================
# STEP 3: GENERATE HTML DASHBOARD
# ============================================================

def generate_html(records, months_order, out_path,
                  overall_avg, low_count, med_count, high_count,
                  tot_forecast, tot_actual):
    print(f"\n[3/4] Generating HTML dashboard...")

    total_assoc = len(records)

    monthly_stats = []
    for m in months_order:
        n   = len([r for r in records if m in r['months']])
        mf  = round(sum(safe_float(r['months'].get(m,{}).get('forecast',0)) for r in records), 1)
        ma  = round(sum(safe_float(r['months'].get(m,{}).get('actual',  0)) for r in records), 1)
        ml  = round(sum(r['months'].get(m,{}).get('leave',0) for r in records), 1)
        mu  = round(sum(r['months'].get(m,{}).get('util',0) for r in records)/n, 1) if n else 0
        monthly_stats.append({'month': m, 'forecast': mf, 'actual': ma, 'leave': ml, 'util': mu})

    acct_data = defaultdict(list)
    for r in records: acct_data[r['Account']].append(r)
    acct_rollup = []
    for acct, recs in sorted(acct_data.items(), key=lambda x: -len(x[1])):
        n     = len(recs)
        avg_u = round(sum(avg_util(r, months_order) for r in recs)/n, 1)
        month_utils = {m: round(sum(r['months'].get(m,{}).get('util',0) for r in recs)/n, 1) for m in months_order}
        acct_rollup.append({'account': acct, 'count': n, 'avg_util': avg_u, 'months': month_utils})

    grade_data = defaultdict(list)
    for r in records: grade_data[r['Grade']].append(r)
    grade_order = ['A','SA','M','SM','AD','PA','PAT','Cont']
    all_grades  = grade_order + [g for g in grade_data if g not in grade_order]
    grade_rollup = [{'grade': g, 'count': len(grade_data[g]),
                     'avg_util': round(sum(avg_util(r, months_order) for r in grade_data[g])/len(grade_data[g]),1)}
                    for g in all_grades if g in grade_data]

    country_data = defaultdict(list)
    for r in records: country_data[r['Country']].append(r)
    country_rollup = [{'country': c, 'count': len(recs),
                       'avg_util': round(sum(avg_util(r, months_order) for r in recs)/len(recs),1)}
                      for c, recs in sorted(country_data.items(), key=lambda x: -len(x[1]))]

    low_emp = sorted([r for r in records if avg_util(r, months_order) < UTIL_HIGH],
                     key=lambda r: avg_util(r, months_order))[:20]

    def badge(u):
        cls = "badge-green" if u >= UTIL_HIGH else "badge-yellow" if u >= UTIL_MEDIUM else "badge-red"
        return f'<span class="badge {cls}">{u}%</span>'

    def td_util(u):
        c = "#dcfce7" if u >= UTIL_HIGH else "#fef9c3" if u >= UTIL_MEDIUM else "#fee2e2"
        return f'<td style="background:{c};text-align:center;font-weight:700;">{u}%</td>'

    acct_rows = ""
    for i, a in enumerate(acct_rollup):
        mcells = "".join(td_util(a['months'][m]) for m in months_order)
        alrt   = "badge-green" if a['avg_util'] >= UTIL_HIGH else "badge-yellow" if a['avg_util'] >= UTIL_MEDIUM else "badge-red"
        acct_rows += f"<tr><td>{i+1}</td><td style='font-weight:600'>{a['account']}</td><td style='text-align:center'>{a['count']}</td><td style='text-align:center'>{badge(a['avg_util'])}</td>{mcells}</tr>"

    low_rows = ""
    for i, r in enumerate(low_emp):
        avg_u  = avg_util(r, months_order)
        mcells = "".join(td_util(r['months'].get(m,{}).get('util',0)) for m in months_order)
        low_rows += f"<tr><td>{i+1}</td><td style='font-weight:600'>{r['Associate Name']}</td><td style='text-align:center'>{r['Grade']}</td><td>{r['Account'][:30]}</td><td style='text-align:center'>{r['Country']}</td><td style='text-align:center'>{badge(avg_u)}</td>{mcells}</tr>"

    monthly_rows = ""
    for ms in monthly_stats:
        c = "#dcfce7" if ms['util'] >= UTIL_HIGH else "#fef9c3" if ms['util'] >= UTIL_MEDIUM else "#fee2e2"
        monthly_rows += f"<tr><td style='font-weight:700'>{ms['month']}</td><td style='text-align:center'>{WORKING_DAYS.get(ms['month'],'-')}</td><td style='text-align:center'>{ms['forecast']}</td><td style='text-align:center'>{ms['actual']}</td><td style='text-align:center'>{ms['leave']}</td><td style='background:{c};text-align:center;font-weight:700'>{ms['util']}%</td></tr>"

    month_headers = "".join(f"<th>{m}</th>" for m in months_order)
    today = date.today().strftime("%d %B %Y")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>QEA UHG Utilization Dashboard 2026</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',sans-serif;background:#f0f4f8;color:#1e293b}}
.header{{background:linear-gradient(135deg,#1e3a5f,#2e75b6);color:#fff;padding:28px 36px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 4px 12px rgba(0,0,0,.2)}}
.header h1{{font-size:1.55rem;font-weight:700}}
.header p{{font-size:.85rem;opacity:.85;margin-top:4px}}
.header-meta{{text-align:right;font-size:.78rem;opacity:.8}}
.container{{max-width:1400px;margin:0 auto;padding:24px}}
.section-title{{font-size:1rem;font-weight:700;color:#1e3a5f;margin:28px 0 14px;padding-left:12px;border-left:4px solid #2e75b6}}
.kpi-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:14px;margin-bottom:8px}}
.kpi-card{{background:#fff;border-radius:12px;padding:18px 14px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.07);border-top:4px solid #2e75b6;transition:transform .2s}}
.kpi-card:hover{{transform:translateY(-3px)}}
.kpi-card .val{{font-size:1.8rem;font-weight:800;color:#1e3a5f}}
.kpi-card .lbl{{font-size:.7rem;color:#64748b;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}}
.kpi-card.green{{border-top-color:#22c55e}}.kpi-card.green .val{{color:#16a34a}}
.kpi-card.yellow{{border-top-color:#f59e0b}}.kpi-card.yellow .val{{color:#d97706}}
.kpi-card.red{{border-top-color:#ef4444}}.kpi-card.red .val{{color:#ef4444}}
.charts-grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:8px}}
.charts-grid-3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px;margin-bottom:8px}}
.chart-card{{background:#fff;border-radius:12px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.07)}}
.chart-card h3{{font-size:.88rem;font-weight:700;color:#1e3a5f;margin-bottom:14px}}
.chart-wrap{{position:relative;height:240px}}
.table-card{{background:#fff;border-radius:12px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.07);margin-bottom:20px;overflow-x:auto}}
.table-card h3{{font-size:.92rem;font-weight:700;color:#1e3a5f;margin-bottom:14px}}
table{{width:100%;border-collapse:collapse;font-size:.81rem}}
th{{background:#1e3a5f;color:#fff;padding:10px;text-align:left;font-size:.77rem;white-space:nowrap}}
td{{padding:8px 10px;border-bottom:1px solid #e2e8f0;white-space:nowrap}}
tr:hover td{{background:#f8fafc}}
tr:last-child td{{border-bottom:none}}
.badge{{display:inline-block;padding:3px 10px;border-radius:20px;font-size:.74rem;font-weight:700}}
.badge-green{{background:#dcfce7;color:#16a34a}}
.badge-yellow{{background:#fef9c3;color:#b45309}}
.badge-red{{background:#fee2e2;color:#dc2626}}
.legend{{display:flex;gap:16px;margin-bottom:14px;flex-wrap:wrap}}
.legend-item{{display:flex;align-items:center;gap:6px;font-size:.77rem}}
.legend-dot{{width:12px;height:12px;border-radius:50%}}
.footer{{text-align:center;padding:20px;color:#94a3b8;font-size:.77rem}}
</style>
</head>
<body>
<div class="header">
  <div>
    <h1>QEA &ndash; UHG Leave &amp; Utilization Dashboard 2026</h1>
    <p>H1 2026 &nbsp;|&nbsp; {" &ndash; ".join([months_order[0], months_order[-1]])} &nbsp;|&nbsp; QEA HC &amp; QEA NFT Pools</p>
  </div>
  <div class="header-meta">
    <div>Generated: {today}</div>
    <div style="margin-top:4px">Utilization = (Working Days &minus; Leave) &divide; Working Days &times; 100</div>
    <div style="margin-top:4px">Leave Used = Actuals where available, else Forecast</div>
  </div>
</div>
<div class="container">
  <div class="section-title">Key Performance Indicators</div>
  <div class="kpi-grid">
    <div class="kpi-card"><div class="val">{total_assoc}</div><div class="lbl">Total Associates</div></div>
    <div class="kpi-card"><div class="val">{int(tot_forecast)}</div><div class="lbl">H1 Forecast Leave (Days)</div></div>
    <div class="kpi-card"><div class="val">{int(tot_actual)}</div><div class="lbl">H1 Actual Leave (Days)</div></div>
    <div class="kpi-card" style="border-top-color:#2e75b6"><div class="val" style="color:#2e75b6">{overall_avg}%</div><div class="lbl">Avg H1 Utilization</div></div>
    <div class="kpi-card green"><div class="val">{high_count}</div><div class="lbl">High Util (&ge;{UTIL_HIGH}%)</div></div>
    <div class="kpi-card yellow"><div class="val">{med_count}</div><div class="lbl">Medium Util ({UTIL_MEDIUM}&ndash;{UTIL_HIGH-1}%)</div></div>
    <div class="kpi-card red"><div class="val">{low_count}</div><div class="lbl">Low Util (&lt;{UTIL_MEDIUM}%)</div></div>
  </div>

  <div class="section-title">Monthly Trends</div>
  <div class="charts-grid">
    <div class="chart-card"><h3>Monthly Avg Utilization % &mdash; H1 2026</h3><div class="chart-wrap"><canvas id="chartUtil"></canvas></div></div>
    <div class="chart-card"><h3>Leave Forecast vs Actuals by Month</h3><div class="chart-wrap"><canvas id="chartLeave"></canvas></div></div>
  </div>

  <div class="section-title">Workforce Breakdown</div>
  <div class="charts-grid-3">
    <div class="chart-card"><h3>Top 10 Accounts &mdash; Avg Utilization %</h3><div class="chart-wrap"><canvas id="chartAcct"></canvas></div></div>
    <div class="chart-card"><h3>Associates by Grade</h3><div class="chart-wrap"><canvas id="chartGrade"></canvas></div></div>
    <div class="chart-card"><h3>Associates by Country</h3><div class="chart-wrap"><canvas id="chartCountry"></canvas></div></div>
  </div>

  <div class="section-title">Monthly Summary</div>
  <div class="table-card">
    <table><thead><tr><th>Month</th><th>Working Days</th><th>Forecast (Days)</th><th>Actuals (Days)</th><th>Leave Used</th><th>Avg Util %</th></tr></thead>
    <tbody>{monthly_rows}</tbody></table>
  </div>

  <div class="section-title">Account-Level Utilization Rollup</div>
  <div class="table-card">
    <table><thead><tr><th>#</th><th>Account</th><th>Associates</th><th>Avg H1 Util%</th>{month_headers}</tr></thead>
    <tbody>{acct_rows}</tbody></table>
  </div>

  <div class="section-title">Utilization Attention List (Top 20 Lowest)</div>
  <div class="table-card">
    <div class="legend">
      <div class="legend-item"><div class="legend-dot" style="background:#22c55e"></div>High (&ge;{UTIL_HIGH}%)</div>
      <div class="legend-item"><div class="legend-dot" style="background:#f59e0b"></div>Medium ({UTIL_MEDIUM}&ndash;{UTIL_HIGH-1}%)</div>
      <div class="legend-item"><div class="legend-dot" style="background:#ef4444"></div>Low (&lt;{UTIL_MEDIUM}%)</div>
    </div>
    <table><thead><tr><th>#</th><th>Associate Name</th><th>Grade</th><th>Account</th><th>Country</th><th>Avg Util%</th>{month_headers}</tr></thead>
    <tbody>{low_rows}</tbody></table>
  </div>
</div>
<div class="footer">QEA &ndash; UHG Leave &amp; Utilization Report &nbsp;&bull;&nbsp; Generated on {today}</div>

<script>
const MONTHS   = {[f'"{m}"' for m in [ms["month"] for ms in monthly_stats]]};
const UTIL     = {[ms["util"] for ms in monthly_stats]};
const FORECAST = {[ms["forecast"] for ms in monthly_stats]};
const ACTUAL   = {[ms["actual"] for ms in monthly_stats]};
const ACCT_LBL = {[f'"{a["account"][:25]}"' for a in acct_rollup[:10]]};
const ACCT_UTL = {[a["avg_util"] for a in acct_rollup[:10]]};
const GRADE_LBL= {[f'"{g["grade"]}"' for g in grade_rollup]};
const GRADE_CNT= {[g["count"] for g in grade_rollup]};
const CTR_LBL  = {[f'"{c["country"]}"' for c in country_rollup]};
const CTR_CNT  = {[c["count"] for c in country_rollup]};

new Chart(document.getElementById('chartUtil'),{{
  type:'line',
  data:{{labels:MONTHS,datasets:[{{label:'Avg Util %',data:UTIL,borderColor:'#2e75b6',backgroundColor:'rgba(46,117,182,.12)',fill:true,tension:0.4,pointRadius:5,borderWidth:2.5}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:true}}}},scales:{{x:{{grid:{{display:false}}}},y:{{min:80,max:100,ticks:{{callback:v=>v+'%'}}}}}}}}
}});

new Chart(document.getElementById('chartLeave'),{{
  type:'bar',
  data:{{labels:MONTHS,datasets:[{{label:'Forecast',data:FORECAST,backgroundColor:'rgba(46,117,182,.75)',borderRadius:4}},{{label:'Actuals',data:ACTUAL,backgroundColor:'rgba(34,197,94,.75)',borderRadius:4}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:true,position:'top'}}}},scales:{{x:{{grid:{{display:false}}}}}}}}
}});

new Chart(document.getElementById('chartAcct'),{{
  type:'bar',
  data:{{labels:ACCT_LBL,datasets:[{{data:ACCT_UTL,backgroundColor:ACCT_UTL.map(u=>u>={UTIL_HIGH}?'rgba(34,197,94,.8)':u>={UTIL_MEDIUM}?'rgba(245,158,11,.8)':'rgba(239,68,68,.8)'),borderRadius:4}}]}},
  options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{x:{{min:80,max:100,ticks:{{callback:v=>v+'%'}}}},y:{{ticks:{{font:{{size:10}}}}}}}}}}
}});

new Chart(document.getElementById('chartGrade'),{{
  type:'doughnut',
  data:{{labels:GRADE_LBL,datasets:[{{data:GRADE_CNT,backgroundColor:['#2e75b6','#70ad47','#ffc000','#ed7d31','#4472c4','#a5a5a5','#5b9bd5','#c00000'],borderWidth:2,borderColor:'#fff'}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'right',labels:{{font:{{size:10}},padding:8}}}}}}}}
}});

new Chart(document.getElementById('chartCountry'),{{
  type:'doughnut',
  data:{{labels:CTR_LBL,datasets:[{{data:CTR_CNT,backgroundColor:['#2e75b6','#ed7d31','#70ad47','#ffc000'],borderWidth:2,borderColor:'#fff'}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'right',labels:{{font:{{size:10}},padding:8}}}}}}}}
}});
</script>
</body></html>"""

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  -> HTML saved: {out_path}")

# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print("  QEA – UHG Leave & Utilization Report Generator")
    print("=" * 60)

    # Get source file path
    if len(sys.argv) > 1:
        src_path = sys.argv[1].strip('"').strip("'")
    else:
        src_path = input("\nEnter path to source Excel file:\n> ").strip().strip('"').strip("'")

    if not os.path.exists(src_path):
        print(f"\n[ERROR] File not found: {src_path}")
        sys.exit(1)

    # Output folder = same as source file
    out_dir   = os.path.dirname(os.path.abspath(src_path))
    today_str = date.today().strftime("%Y-%m-%d")
    xlsx_out  = os.path.join(out_dir, f"QEA-UHG-Utilization-Report-{today_str}.xlsx")
    html_out  = os.path.join(out_dir, f"QEA-UHG-Utilization-Dashboard-{today_str}.html")

    # Run pipeline
    records, months_order = read_source(src_path)

    stats = generate_excel(records, months_order, xlsx_out)
    avg_utils, overall_avg, low_count, med_count, high_count, tot_forecast, tot_actual = stats

    generate_html(records, months_order, html_out,
                  overall_avg, low_count, med_count, high_count,
                  tot_forecast, tot_actual)

    print(f"\n[4/4] Done!")
    print(f"\n  Excel  : {xlsx_out}")
    print(f"  HTML   : {html_out}")
    print(f"\n  Associates : {len(records)}")
    print(f"  Avg H1 Util: {overall_avg}%")
    print(f"  High (>=90%): {high_count}  |  Medium (80-90%): {med_count}  |  Low (<80%): {low_count}")

    # Auto-open HTML in browser
    try:
        import subprocess
        subprocess.Popen(['start', '', html_out], shell=True)
    except:
        pass

if __name__ == "__main__":
    main()
