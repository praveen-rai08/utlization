import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Working days per month 2026
working_days = {
    "Jan'26": 21, "Feb'26": 20, "Mar'26": 22,
    "Apr'26": 22, "May'26": 21, "Jun'26": 22,
}
months_order = list(working_days.keys())

# Colour palette
C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "BDD7EE"
C_VERY_LIGHT = "DEEAF1"
C_GREEN      = "70AD47"
C_LIGHT_GREEN= "E2EFDA"
C_YELLOW     = "FFD966"
C_LIGHT_YEL  = "FFF2CC"
C_RED        = "FF0000"
C_LIGHT_RED  = "FFE0E0"
C_WHITE      = "FFFFFF"
C_GREY       = "F2F2F2"
C_DARK_GREY  = "595959"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_DARK_GREY, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="595959")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# Read source data
wb_src = openpyxl.load_workbook(
    r'C:\Users\202294\Downloads\QEA-UHG-Leave-Forecast-2026.xlsx',
    data_only=True
)

sheet_months = [
    ('2026-Jan-Feb-Mar', [("Jan'26", 14, 15), ("Feb'26", 18, 19), ("Mar'26", 22, 23)]),
    ('2026-Apr-May-Jun', [("Apr'26", 14, 15), ("May'26", 18, 19), ("Jun'26", 22, 23)]),
]

all_records = {}

for sheet_name, month_configs in sheet_months:
    ws = wb_src[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[2:]:
        if not any(row):
            continue
        assoc_id = row[1]
        if not assoc_id:
            continue
        if assoc_id not in all_records:
            all_records[assoc_id] = {
                'Associate ID': assoc_id,
                'Associate Name': row[2] or '',
                'Grade': row[3] or '',
                'Account': row[7] or '',
                'Project': row[5] or '',
                'Billability': row[10] or '',
                'Country': row[11] or '',
                'Onsite/Offshore': row[12] or '',
                'City': row[13] or '',
                'months': {}
            }
        for month_name, fc_idx, act_idx in month_configs:
            forecast = row[fc_idx]
            actual   = row[act_idx]
            wd       = working_days[month_name]
            leave    = actual if actual is not None else (forecast if forecast is not None else 0)
            try:
                leave = float(leave)
            except:
                leave = 0.0
            available = wd - leave
            util_pct  = round((available / wd) * 100, 1) if wd > 0 else 0.0
            all_records[assoc_id]['months'][month_name] = {
                'wd': wd,
                'forecast': forecast if forecast is not None else 0,
                'actual': actual,
                'leave': leave,
                'available': available,
                'util': util_pct,
            }

records = list(all_records.values())

def avg_util(rec):
    vals = [rec['months'][m]['util'] for m in months_order if m in rec['months']]
    return round(sum(vals)/len(vals), 1) if vals else 0

# Create output workbook
wb = openpyxl.Workbook()

# ============================================================
# SHEET 1 - SUMMARY DASHBOARD
# ============================================================
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_view.showGridLines = False

ws_dash.merge_cells("A1:N1")
ws_dash["A1"] = "QEA - UHG Leave & Utilization Report 2026 (H1)"
ws_dash["A1"].font      = Font(bold=True, size=16, color=C_WHITE)
ws_dash["A1"].fill      = fill(C_DARK_BLUE)
ws_dash["A1"].alignment = center()
ws_dash.row_dimensions[1].height = 36

ws_dash.merge_cells("A2:N2")
ws_dash["A2"] = "Utilization % = (Working Days - Leave Days) / Working Days x 100   |   Leave Used = Actuals where available, else Forecast"
ws_dash["A2"].font      = Font(italic=True, size=9, color=C_WHITE)
ws_dash["A2"].fill      = fill(C_MID_BLUE)
ws_dash["A2"].alignment = center()
ws_dash.row_dimensions[2].height = 18

# KPI stats
total_assoc    = len(records)
def safe_float(v, default=0):
    try:
        return float(v) if v is not None else default
    except:
        return default

total_forecast = sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) for r in records for m in months_order)
total_actual   = sum(safe_float(r['months'].get(m, {}).get('actual', 0)) for r in records for m in months_order)
avg_utils      = [avg_util(r) for r in records]
overall_avg    = round(sum(avg_utils)/len(avg_utils), 1) if avg_utils else 0
low_count      = sum(1 for u in avg_utils if u < 80)
med_count      = sum(1 for u in avg_utils if 80 <= u < 90)
high_count     = sum(1 for u in avg_utils if u >= 90)

kpi_labels = ["Total Associates", "Total H1 Forecast (Days)", "Total H1 Actuals (Days)",
              "Avg H1 Utilization %", "Low Util (<80%)", "Medium Util (80-90%)", "High Util (>=90%)"]
kpi_vals   = [total_assoc, int(total_forecast), int(total_actual),
              str(overall_avg)+"%", low_count, med_count, high_count]
kpi_fills  = [C_MID_BLUE, C_MID_BLUE, C_MID_BLUE, C_MID_BLUE,
              "FFB3B3", "FFF0B3", "C6EFCE"]
kpi_col_pairs = [(1,2),(3,4),(5,6),(7,8),(9,10),(11,12),(13,14)]

kpi_start_row = 4
for i, (label, val, kfill, (c1, c2)) in enumerate(zip(kpi_labels, kpi_vals, kpi_fills, kpi_col_pairs)):
    lr = kpi_start_row
    vr = kpi_start_row + 1
    lc1 = get_column_letter(c1)
    lc2 = get_column_letter(c2)
    ws_dash.merge_cells(f"{lc1}{lr}:{lc2}{lr}")
    ws_dash.merge_cells(f"{lc1}{vr}:{lc2}{vr}")
    ws_dash[f"{lc1}{lr}"] = label
    ws_dash[f"{lc1}{lr}"].font      = Font(bold=True, size=9, color=C_DARK_GREY)
    ws_dash[f"{lc1}{lr}"].fill      = fill(C_GREY)
    ws_dash[f"{lc1}{lr}"].alignment = center()
    ws_dash[f"{lc1}{vr}"] = val
    ws_dash[f"{lc1}{vr}"].font      = Font(bold=True, size=14, color=C_DARK_BLUE)
    ws_dash[f"{lc1}{vr}"].fill      = fill(kfill)
    ws_dash[f"{lc1}{vr}"].alignment = center()
    ws_dash.row_dimensions[lr].height = 20
    ws_dash.row_dimensions[vr].height = 32

# Monthly summary table
ms_row = 8
ws_dash.merge_cells(f"A{ms_row}:H{ms_row}")
ws_dash[f"A{ms_row}"] = "Monthly Leave & Utilization Summary"
ws_dash[f"A{ms_row}"].font      = Font(bold=True, size=11, color=C_WHITE)
ws_dash[f"A{ms_row}"].fill      = fill(C_DARK_BLUE)
ws_dash[f"A{ms_row}"].alignment = center()
ws_dash.row_dimensions[ms_row].height = 22

hdr_cols = ["Month", "Working Days", "Total Forecast (Days)", "Total Actuals (Days)",
            "Total Leave Used", "Total Available Days", "Avg Util %", "Associates"]
for ci, h in enumerate(hdr_cols, 1):
    c = ws_dash.cell(ms_row+1, ci, h)
    c.font      = Font(bold=True, size=9, color=C_WHITE)
    c.fill      = fill(C_MID_BLUE)
    c.alignment = center()
    c.border    = border_thin()
ws_dash.row_dimensions[ms_row+1].height = 22

for ri, m in enumerate(months_order, ms_row+2):
    wd = working_days[m]
    mf = sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) for r in records)
    ma = sum(safe_float(r['months'].get(m, {}).get('actual', 0)) for r in records)
    ml = sum(r['months'].get(m, {}).get('leave', 0) for r in records)
    mav= sum(r['months'].get(m, {}).get('available', 0) for r in records)
    mu = round(sum(r['months'].get(m, {}).get('util', 0) for r in records)/len(records), 1)
    mc = sum(1 for r in records if m in r['months'])
    row_data = [m, wd, round(mf,1), round(ma,1), round(ml,1), round(mav,1), str(mu)+"%", mc]
    rf = C_VERY_LIGHT if (ri % 2 == 0) else C_WHITE
    for ci, val in enumerate(row_data, 1):
        c = ws_dash.cell(ri, ci, val)
        c.fill      = fill(rf)
        c.font      = font(size=9)
        c.alignment = center()
        c.border    = border_thin()
    ws_dash.row_dimensions[ri].height = 18

for col, w in zip("ABCDEFGHIJKLMN", [14,14,18,18,16,16,12,12,10,10,10,10,10,10]):
    ws_dash.column_dimensions[col].width = w

# ============================================================
# SHEET 2 - EMPLOYEE UTILIZATION DETAIL
# ============================================================
ws_emp = wb.create_sheet("Employee Utilization")
ws_emp.sheet_view.showGridLines = False
ws_emp.freeze_panes = "J3"

ws_emp.merge_cells("A1:AH1")
ws_emp["A1"] = "Employee-Level Leave & Utilization Detail - H1 2026"
ws_emp["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
ws_emp["A1"].fill      = fill(C_DARK_BLUE)
ws_emp["A1"].alignment = center()
ws_emp.row_dimensions[1].height = 30

static_hdrs = ["S.No","Associate ID","Associate Name","Grade","Account",
               "Project","Billability","Country","Onsite/Offshore","City",
               "Avg H1 Util %","Alert"]
n_static = len(static_hdrs)

for ci, h in enumerate(static_hdrs, 1):
    c = ws_emp.cell(2, ci, h)
    c.font      = Font(bold=True, size=9, color=C_WHITE)
    c.fill      = fill(C_DARK_BLUE)
    c.alignment = center()
    c.border    = border_thin()

month_col_fills = ["2E75B6","1B6CA8","155A8A","0F4A72","0A3A5A","052A42"]
month_sub_hdrs  = ["Forecast\n(Days)", "Actuals\n(Days)", "Leave\nUsed", "Available\nDays", "Util %"]

for mi, m in enumerate(months_order):
    start_ci = n_static + 1 + mi * len(month_sub_hdrs)
    end_ci   = start_ci + len(month_sub_hdrs) - 1
    sc = get_column_letter(start_ci)
    ec = get_column_letter(end_ci)
    ws_emp.merge_cells(f"{sc}2:{ec}2")
    ws_emp[f"{sc}2"] = m
    ws_emp[f"{sc}2"].font      = Font(bold=True, size=9, color=C_WHITE)
    ws_emp[f"{sc}2"].fill      = fill(month_col_fills[mi % len(month_col_fills)])
    ws_emp[f"{sc}2"].alignment = center()
    ws_emp[f"{sc}2"].border    = border_medium()

for mi, m in enumerate(months_order):
    for si, sh in enumerate(month_sub_hdrs):
        ci = n_static + 1 + mi * len(month_sub_hdrs) + si
        c  = ws_emp.cell(3, ci, sh)
        c.font      = Font(bold=True, size=8, color=C_WHITE)
        c.fill      = fill(month_col_fills[mi % len(month_col_fills)])
        c.alignment = center()
        c.border    = border_thin()

for ci in range(1, n_static+1):
    c = ws_emp.cell(3, ci, "")
    c.fill   = fill(C_DARK_BLUE)
    c.border = border_thin()

ws_emp.row_dimensions[2].height = 22
ws_emp.row_dimensions[3].height = 30

records_sorted = sorted(records, key=lambda r: r['Associate Name'])

for ri, rec in enumerate(records_sorted, 4):
    avg_u = avg_util(rec)
    if avg_u < 80:
        alert, row_fill = "Low (<80%)", "FFE0E0"
    elif avg_u < 90:
        alert, row_fill = "Medium (80-90%)", "FFF2CC"
    else:
        alert, row_fill = "High (>=90%)", "E2EFDA"

    static_vals = [ri-3, rec['Associate ID'], rec['Associate Name'], rec['Grade'],
                   rec['Account'], rec['Project'], rec['Billability'],
                   rec['Country'], rec['Onsite/Offshore'], rec['City'],
                   str(avg_u)+"%", alert]

    for ci, val in enumerate(static_vals, 1):
        c = ws_emp.cell(ri, ci, val)
        c.fill      = fill(row_fill)
        c.font      = font(size=9, bold=(ci == 3))
        c.alignment = left() if ci == 3 else center()
        c.border    = border_thin()

    for mi, m in enumerate(months_order):
        md   = rec['months'].get(m, {})
        vals = [
            md.get('forecast', ''),
            md.get('actual', ''),
            md.get('leave', ''),
            md.get('available', ''),
            str(md.get('util', ''))+"%"  if md.get('util', '') != '' else ''
        ]
        for si, val in enumerate(vals):
            ci = n_static + 1 + mi * len(month_sub_hdrs) + si
            c  = ws_emp.cell(ri, ci, val)
            if si == 4 and md.get('util', '') != '':
                u = md['util']
                if u < 80:   cf = "FFB3B3"
                elif u < 90: cf = "FFF0B3"
                else:        cf = "C6EFCE"
                c.fill = fill(cf)
            else:
                c.fill = fill(row_fill)
            c.font      = font(size=9, bold=(si == 4))
            c.alignment = center()
            c.border    = border_thin()

    ws_emp.row_dimensions[ri].height = 16

emp_widths = [5, 12, 24, 7, 22, 28, 10, 14, 12, 12, 10, 16]
for ci, w in enumerate(emp_widths, 1):
    ws_emp.column_dimensions[get_column_letter(ci)].width = w
for mi in range(len(months_order)):
    for si in range(len(month_sub_hdrs)):
        ci = n_static + 1 + mi * len(month_sub_hdrs) + si
        ws_emp.column_dimensions[get_column_letter(ci)].width = 10

# ============================================================
# SHEET 3 - LOW UTILIZATION ALERTS
# ============================================================
ws_alert = wb.create_sheet("Low Util Alerts")
ws_alert.sheet_view.showGridLines = False
ws_alert.freeze_panes = "A4"

ws_alert.merge_cells("A1:M1")
ws_alert["A1"] = "Low Utilization Alert Report - Associates with Avg H1 Util < 80%"
ws_alert["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
ws_alert["A1"].fill      = fill("C00000")
ws_alert["A1"].alignment = center()
ws_alert.row_dimensions[1].height = 30

ws_alert.merge_cells("A2:M2")
ws_alert["A2"] = "These associates have average H1 utilization below 80% - review leave patterns and plan accordingly."
ws_alert["A2"].font      = Font(italic=True, size=9, color="C00000")
ws_alert["A2"].fill      = fill("FFE0E0")
ws_alert["A2"].alignment = center()
ws_alert.row_dimensions[2].height = 18

alert_hdrs = ["S.No","Associate ID","Associate Name","Grade","Account","Country",
              "Onsite/Offshore","Avg Util %"] + [f"{m}\nUtil %" for m in months_order]

for ci, h in enumerate(alert_hdrs, 1):
    c = ws_alert.cell(3, ci, h)
    c.font      = Font(bold=True, size=9, color=C_WHITE)
    c.fill      = fill("C00000")
    c.alignment = center()
    c.border    = border_thin()
ws_alert.row_dimensions[3].height = 28

low_util_recs = sorted(
    [r for r in records if avg_util(r) < 80],
    key=lambda r: avg_util(r)
)

for ri, rec in enumerate(low_util_recs, 4):
    avg_u    = avg_util(rec)
    row_fill = "FFD7D7" if avg_u < 70 else "FFE0E0"
    base_vals = [ri-3, rec['Associate ID'], rec['Associate Name'], rec['Grade'],
                 rec['Account'], rec['Country'], rec['Onsite/Offshore'], str(avg_u)+"%"]
    month_utils = []
    for m in months_order:
        u = rec['months'].get(m, {}).get('util', '')
        month_utils.append(str(u)+"%" if u != '' else '')
    all_vals = base_vals + month_utils
    for ci, val in enumerate(all_vals, 1):
        c = ws_alert.cell(ri, ci, val)
        c.fill      = fill(row_fill)
        c.font      = font(size=9, bold=(ci == 8))
        c.alignment = left() if ci == 3 else center()
        c.border    = border_thin()
        if ci >= 9:
            try:
                u_val = float(str(val).replace('%', ''))
                if u_val < 80: c.fill = fill("FFB3B3")
            except:
                pass
    ws_alert.row_dimensions[ri].height = 16

alert_widths = [5, 12, 24, 7, 22, 14, 12, 10, 10, 10, 10, 10, 10]
for ci, w in enumerate(alert_widths, 1):
    ws_alert.column_dimensions[get_column_letter(ci)].width = w

# ============================================================
# SHEET 4 - ACCOUNT-LEVEL ROLLUP
# ============================================================
ws_acct = wb.create_sheet("Account Rollup")
ws_acct.sheet_view.showGridLines = False
ws_acct.freeze_panes = "A3"

ws_acct.merge_cells("A1:P1")
ws_acct["A1"] = "Account-Level Leave & Utilization Rollup - H1 2026"
ws_acct["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
ws_acct["A1"].fill      = fill(C_DARK_BLUE)
ws_acct["A1"].alignment = center()
ws_acct.row_dimensions[1].height = 30

acct_hdrs = ["S.No","Account Name","# Associates","Avg H1 Util %","Alert Level"]
for m in months_order:
    acct_hdrs += [f"{m}\nForecast", f"{m}\nActuals", f"{m}\nAvg Util %"]

for ci, h in enumerate(acct_hdrs, 1):
    c = ws_acct.cell(2, ci, h)
    c.font      = Font(bold=True, size=9, color=C_WHITE)
    c.fill      = fill(C_DARK_BLUE)
    c.alignment = center()
    c.border    = border_thin()
ws_acct.row_dimensions[2].height = 30

acct_data   = defaultdict(list)
for r in records:
    acct_data[r['Account']].append(r)

for ri, (acct, recs) in enumerate(sorted(acct_data.items(), key=lambda x: -len(x[1])), 3):
    n     = len(recs)
    avg_u = round(sum(avg_util(r) for r in recs)/n, 1)
    if avg_u < 80:   alert, rf = "Low",    "FFE0E0"
    elif avg_u < 90: alert, rf = "Medium", "FFF2CC"
    else:            alert, rf = "High",   "E2EFDA"

    base = [ri-2, acct, n, str(avg_u)+"%", alert]
    for ci, val in enumerate(base, 1):
        c = ws_acct.cell(ri, ci, val)
        c.fill      = fill(rf)
        c.font      = font(size=9, bold=(ci in [2, 4]))
        c.alignment = left() if ci == 2 else center()
        c.border    = border_thin()

    for mi, m in enumerate(months_order):
        mf   = round(sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) for r in recs), 1)
        ma   = round(sum(safe_float(r['months'].get(m, {}).get('actual', 0)) for r in recs), 1)
        mu   = round(sum(r['months'].get(m, {}).get('util', 0) for r in recs)/n, 1)
        ci_base = 6 + mi * 3
        for ci, val in zip([ci_base, ci_base+1, ci_base+2], [mf, ma, str(mu)+"%"]):
            c = ws_acct.cell(ri, ci, val)
            if ci == ci_base+2:
                if mu < 80:   cf = "FFB3B3"
                elif mu < 90: cf = "FFF0B3"
                else:         cf = "C6EFCE"
                c.fill = fill(cf)
            else:
                c.fill = fill(rf)
            c.font      = font(size=9, bold=(ci == ci_base+2))
            c.alignment = center()
            c.border    = border_thin()
    ws_acct.row_dimensions[ri].height = 18

acct_widths = [5, 26, 10, 10, 10]
for ci, w in enumerate(acct_widths, 1):
    ws_acct.column_dimensions[get_column_letter(ci)].width = w
for mi in range(len(months_order)):
    for si in range(3):
        ci = 6 + mi * 3 + si
        ws_acct.column_dimensions[get_column_letter(ci)].width = 11

# ============================================================
# SHEET 5 - GRADE ROLLUP
# ============================================================
ws_grade = wb.create_sheet("Grade Rollup")
ws_grade.sheet_view.showGridLines = False

ws_grade.merge_cells("A1:P1")
ws_grade["A1"] = "Grade-Level Leave & Utilization Rollup - H1 2026"
ws_grade["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
ws_grade["A1"].fill      = fill(C_DARK_BLUE)
ws_grade["A1"].alignment = center()
ws_grade.row_dimensions[1].height = 30

grade_hdrs = ["Grade","# Associates","Avg H1 Util %","Alert Level"]
for m in months_order:
    grade_hdrs += [f"{m}\nForecast", f"{m}\nActuals", f"{m}\nAvg Util %"]

for ci, h in enumerate(grade_hdrs, 1):
    c = ws_grade.cell(2, ci, h)
    c.font      = Font(bold=True, size=9, color=C_WHITE)
    c.fill      = fill(C_DARK_BLUE)
    c.alignment = center()
    c.border    = border_thin()
ws_grade.row_dimensions[2].height = 30

grade_data  = defaultdict(list)
for r in records:
    grade_data[r['Grade']].append(r)

grade_order = ['A','SA','M','SM','AD','PA','PAT','Cont']
all_grades  = grade_order + [g for g in grade_data if g not in grade_order]

for ri, grade in enumerate([g for g in all_grades if g in grade_data], 3):
    recs  = grade_data[grade]
    n     = len(recs)
    avg_u = round(sum(avg_util(r) for r in recs)/n, 1)
    if avg_u < 80:   alert, rf = "Low",    "FFE0E0"
    elif avg_u < 90: alert, rf = "Medium", "FFF2CC"
    else:            alert, rf = "High",   "E2EFDA"

    base = [grade, n, str(avg_u)+"%", alert]
    for ci, val in enumerate(base, 1):
        c = ws_grade.cell(ri, ci, val)
        c.fill      = fill(rf)
        c.font      = font(size=9, bold=True)
        c.alignment = center()
        c.border    = border_thin()

    for mi, m in enumerate(months_order):
        mf   = round(sum(safe_float(r['months'].get(m, {}).get('forecast', 0)) for r in recs), 1)
        ma   = round(sum(safe_float(r['months'].get(m, {}).get('actual', 0)) for r in recs), 1)
        mu   = round(sum(r['months'].get(m, {}).get('util', 0) for r in recs)/n, 1)
        ci_base = 5 + mi * 3
        for ci, val in zip([ci_base, ci_base+1, ci_base+2], [mf, ma, str(mu)+"%"]):
            c = ws_grade.cell(ri, ci, val)
            if ci == ci_base+2:
                if mu < 80:   cf = "FFB3B3"
                elif mu < 90: cf = "FFF0B3"
                else:         cf = "C6EFCE"
                c.fill = fill(cf)
            else:
                c.fill = fill(rf)
            c.font      = font(size=9, bold=(ci == ci_base+2))
            c.alignment = center()
            c.border    = border_thin()
    ws_grade.row_dimensions[ri].height = 18

for ci, w in enumerate([10,12,12,14]+[11]*18, 1):
    ws_grade.column_dimensions[get_column_letter(ci)].width = w

# ============================================================
# SHEET 6 - COUNTRY ROLLUP
# ============================================================
ws_country = wb.create_sheet("Country Rollup")
ws_country.sheet_view.showGridLines = False

ws_country.merge_cells("A1:K1")
ws_country["A1"] = "Country & Onsite/Offshore - Leave & Utilization Rollup - H1 2026"
ws_country["A1"].font      = Font(bold=True, size=13, color=C_WHITE)
ws_country["A1"].fill      = fill(C_DARK_BLUE)
ws_country["A1"].alignment = center()
ws_country.row_dimensions[1].height = 30

country_hdrs = ["Country","Onsite/Offshore","# Associates","Avg H1 Util %","Alert Level"]
for m in months_order:
    country_hdrs.append(f"{m}\nAvg Util %")

for ci, h in enumerate(country_hdrs, 1):
    c = ws_country.cell(2, ci, h)
    c.font      = Font(bold=True, size=9, color=C_WHITE)
    c.fill      = fill(C_DARK_BLUE)
    c.alignment = center()
    c.border    = border_thin()
ws_country.row_dimensions[2].height = 30

country_oo_data = defaultdict(list)
for r in records:
    country_oo_data[(r['Country'], r['Onsite/Offshore'])].append(r)

for ri, ((country, oo), recs) in enumerate(
        sorted(country_oo_data.items(), key=lambda x: -len(x[1])), 3):
    n     = len(recs)
    avg_u = round(sum(avg_util(r) for r in recs)/n, 1)
    if avg_u < 80:   alert, rf = "Low",    "FFE0E0"
    elif avg_u < 90: alert, rf = "Medium", "FFF2CC"
    else:            alert, rf = "High",   "E2EFDA"

    base = [country, oo, n, str(avg_u)+"%", alert]
    for ci, val in enumerate(base, 1):
        c = ws_country.cell(ri, ci, val)
        c.fill      = fill(rf)
        c.font      = font(size=9, bold=True)
        c.alignment = center()
        c.border    = border_thin()

    for mi, m in enumerate(months_order):
        mu = round(sum(r['months'].get(m, {}).get('util', 0) for r in recs)/n, 1)
        ci = 6 + mi
        c  = ws_country.cell(ri, ci, str(mu)+"%")
        if mu < 80:   cf = "FFB3B3"
        elif mu < 90: cf = "FFF0B3"
        else:         cf = "C6EFCE"
        c.fill      = fill(cf)
        c.font      = font(size=9, bold=True)
        c.alignment = center()
        c.border    = border_thin()
    ws_country.row_dimensions[ri].height = 18

for ci, w in enumerate([16,14,12,12,14,12,12,12,12,12,12], 1):
    ws_country.column_dimensions[get_column_letter(ci)].width = w

# Save
out_path = r'C:\Users\202294\Downloads\QEA-UHG-Utilization-Report-2026.xlsx'
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Associates: {len(records)}")
print(f"Low util (<80%):    {low_count}")
print(f"Medium util (80-90%): {med_count}")
print(f"High util (>=90%):  {high_count}")
print(f"Overall avg H1 util: {overall_avg}%")
